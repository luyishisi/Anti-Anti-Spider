# -*- coding: UTF-8 -*-
# 运行方式： python3 test.py
# 运行要求： urllist.txt
import requests
import json
# import xlwt
# 这个模块每个表只能存65535行 已经弃用
import openpyxl
import sys
import math
import time
import random
import urllib2
import base64
import MySQLdb
import requesocks

data={'sk':'super', 'ak':'super', 'app_id':'1553059213389964', '_skip':'0', '_limit':'99999', 'status':'in:publish,published' }
head={'Accept':'application/json', 'Accept-Encoding':'gzip, deflate, sdch', 'Host':'baijiahao.baidu.com', 'X-Requested-With':'XMLHttpRequest' }

reload(sys)
sys.setdefaultencoding('utf-8')
Type = sys.getfilesystemencoding()

Table = "baijiahao"#sys.argv[1]

#HOST, USER, PASSWD, DB, PORT = '127.0.0.1','root','luyi123', 'my_db', 3306#需要修改
HOST, USER, PASSWD, DB, PORT = '','','', '', 3306#需要修改


select_sql = "SELECT * FROM %s "
#into_sql = "INSERT into baijiahao (title,url,classify,comment_amount,read_amount,newstime) VALUES ('%s','%s','%s',%s,%s,'%s');"
into_sql = "INSERT into "+Table+" (title,url,classify,comment_amount,read_amount,newstime) VALUES (%s,%s,%s,%s,%s,%s);"

def ConnectDB():
    "Connect MySQLdb and Print version."
    connect, cursor = None, None
    count = 0
    while True:
        try :
            connect = MySQLdb.connect(host=HOST, user=USER, passwd=PASSWD, db=DB, port = PORT, charset ='utf8')
            cursor = connect.cursor()
            break
        except MySQLdb.Error, e:
            print "Error %d: %s" % (e.args[0],e.args[1])
            count += 1
            time.sleep(10)
            if count > 100:
                print 'error > 100 end'
                sys.exit(1)
    return connect, cursor

if __name__ =='__main__':
    print 'begin'
    connect, cursor = ConnectDB()
    file = open('urllist.txt') # urllist.txt在你运行的目录下
    idlist=[]
    for i in file.readlines():
        idlist.append(i)

    # f = xlwt.Workbook()
    # sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok=True)
    wb = openpyxl.Workbook()
    # 创建一个sheet
    ws = wb.create_sheet()
    num = 0;
    date_list = []
    for url in idlist:
        data['app_id']=url
        try:
            req=requests.post('http://baijiahao.baidu.com/api/content/article/listall',data=data,headers=head)
        except:
            continue
            time.sleep(2)
            print('post error')
        try:
            myjson=req.json()
        except:
            print('url error:'),url
            continue
        myjson=req.json()
        myjson=myjson['items']
        # 在items下存的就是我们要的内容
        date_list = []
        for i in myjson:
            newstime=i["publish_at"]
            title=i["title"]
            url=i['url']
            comment_amount=i['comment_amount']
            read_amount=i['read_amount']
            classify = i['domain']
            num += 1
            try:
                date_list.append([title,url,classify,int(comment_amount),int(read_amount),newstime])
                #cursor.execute(into_sql % (title,url,classify,int(comment_amount),int(read_amount),newstime))
            except Exception,e:
                print e
            if num % 500 == 0:
                print num,'commit'
                cursor.executemany(into_sql,date_list)
                connect.commit()
                date_list = []
        #cursor.executemany(into_sql , date_list)
    print 'commit'
    cursor.executemany(into_sql,date_list)
    connect.commit()
    print('save to chen.xlsx')
    #wb.save('chen_500.xlsx')
