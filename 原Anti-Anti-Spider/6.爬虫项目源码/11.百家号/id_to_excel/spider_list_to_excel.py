# -*- coding: UTF-8 -*-
# 运行方式： python3 test.py
# 运行要求： urllist.txt
import requests
import json
# import xlwt
# 这个模块每个表只能存65535行 已经弃用
import openpyxl
# 写入slxs的模块

import time
# import pymysql

data={'sk':'super',
'ak':'super',
'app_id':'1553059213389964',
'_skip':'0',
'_limit':'99999',
# 获取的新闻条数
'status':'in:publish,published'
}

head={'Accept':'application/json',
'Accept-Encoding':'gzip, deflate, sdch',
# Accept-Language:zh-CN,zh;q=0.8
# Connection:keep-alive
# Cookie:BAIDUID=7D49D3DEC1E42F447BA7C8A27E2F6A5F:FG=1; BIDUPSID=7D49D3DEC1E42F447BA7C8A27E2F6A5F; PSTM=1484963567; Hm_lvt_192fa266ff34772c28e4ddb36b8f4472=1484967469,1484990376,1484990690; Hm_lpvt_192fa266ff34772c28e4ddb36b8f4472=1484990690
'Host':'baijiahao.baidu.com',
# Referer:http://baijiahao.baidu.com/u?app_id=1547814148759979
 #User-Agent:Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36
'X-Requested-With':'XMLHttpRequest'
}
# def writetomysql(data):
    # db = pymysql.connect("copie.cn","root",'copyist12copyist')
if __name__ =='__main__':
    file = open('urllist.txt') # urllist.txt在你运行的目录下
    idlist=[]
    for i in file.readlines():
        idlist.append(i)

    # f = xlwt.Workbook()
    # sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok=True)
    wb = openpyxl.Workbook()
    # 创建一个sheet
    ws = wb.create_sheet()

    for url in idlist:
        data['app_id']=url
        try:
            req=requests.post('http://baijiahao.baidu.com/api/content/article/listall',data=data,headers=head)
            # print(req.text)https://baijiahao.baidu.com/api/content/article/listall?sk=super&ak=super&app_id=1548780976740890&_skip=18&_limit=12&status=in:publish,published&_preload_statistic=1&_timg_cover=50,172,1000&_cache=1
        except:
            continue
            time.sleep(2)
            print('post error')
            # 有时候会失败我也不知到为什么   超时什么的
        try:
            myjson=req.json()
        except:
            print('url error')
            # 解释失败就是这个id已经被封le
            continue
        myjson=req.json()
        myjson=myjson['items']
        # 在items下存的就是我们要的内容
        for i in myjson:
            newstime=i["publish_at"]
            title=i["title"]
            url=i['url']
            comment_amount=i['comment_amount']
            read_amount=i['read_amount']
            domain = i['domain']
            #print read_amount,domain.encode('utf-8')
            print newstime.split('-')[0]
            #2017-01-01 16:04:03
            #time.sleep(10)
            if(newstime.split('-')[0] == str('2017')):
                try:
                    ws.append([title,url,domain,comment_amount,read_amount,newstime])
                except:
                    print('write one row error ')
    print('save to linshi.xlsx')
    wb.save('urllist_2_7.xlsx')
